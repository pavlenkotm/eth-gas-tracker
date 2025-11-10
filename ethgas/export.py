"""Data export functionality for CSV and Excel formats."""
import csv
import json
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime


class DataExporter:
    """Export historical gas data to various formats."""

    @staticmethod
    def export_to_csv(records: List[Dict], output_path: str) -> None:
        """
        Export records to CSV format.

        Args:
            records: List of gas data records
            output_path: Path to save CSV file
        """
        if not records:
            raise ValueError("No records to export")

        # Define CSV columns
        fieldnames = [
            "timestamp",
            "network",
            "base_fee",
            "priority_tip",
            "max_fee",
            "token_price_usd"
        ]

        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for record in records:
                # Ensure all fields exist
                row = {field: record.get(field, "") for field in fieldnames}
                writer.writerow(row)

    @staticmethod
    def export_to_excel(records: List[Dict], output_path: str) -> None:
        """
        Export records to Excel format using openpyxl.

        Args:
            records: List of gas data records
            output_path: Path to save Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

        if not records:
            raise ValueError("No records to export")

        wb = Workbook()
        ws = wb.active
        ws.title = "Gas Price History"

        # Define headers
        headers = [
            "Timestamp",
            "Network",
            "Base Fee (Gwei)",
            "Priority Tip (Gwei)",
            "Max Fee (Gwei)",
            "Token Price (USD)"
        ]

        # Style header row
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for row_num, record in enumerate(records, 2):
            ws.cell(row=row_num, column=1, value=record.get("timestamp", ""))
            ws.cell(row=row_num, column=2, value=record.get("network", ""))
            ws.cell(row=row_num, column=3, value=record.get("base_fee", 0))
            ws.cell(row=row_num, column=4, value=record.get("priority_tip", 0))
            ws.cell(row=row_num, column=5, value=record.get("max_fee", 0))
            ws.cell(row=row_num, column=6, value=record.get("token_price_usd", 0))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)

    @staticmethod
    def export_to_json(records: List[Dict], output_path: str) -> None:
        """
        Export records to JSON format.

        Args:
            records: List of gas data records
            output_path: Path to save JSON file
        """
        if not records:
            raise ValueError("No records to export")

        with open(output_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(records, jsonfile, indent=2, ensure_ascii=False)

    @staticmethod
    def export_statistics_to_csv(stats: Dict, output_path: str) -> None:
        """
        Export statistics summary to CSV.

        Args:
            stats: Statistics dictionary
            output_path: Path to save CSV file
        """
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Metric", "Value"])

            for key, value in stats.items():
                if isinstance(value, dict):
                    writer.writerow([key, ""])
                    for sub_key, sub_value in value.items():
                        writer.writerow([f"  {sub_key}", sub_value])
                else:
                    writer.writerow([key, value])

    @staticmethod
    def auto_export(records: List[Dict],
                   format: str = "csv",
                   filename: Optional[str] = None) -> str:
        """
        Automatically export with timestamp-based filename.

        Args:
            records: List of gas data records
            format: Export format ('csv', 'excel', or 'json')
            filename: Optional custom filename (default: auto-generated)

        Returns:
            Path to exported file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ethgas_export_{timestamp}"

        # Ensure proper extension
        extensions = {
            "csv": ".csv",
            "excel": ".xlsx",
            "json": ".json"
        }
        ext = extensions.get(format, ".csv")
        if not filename.endswith(ext):
            filename += ext

        # Export to appropriate format
        if format == "csv":
            DataExporter.export_to_csv(records, filename)
        elif format == "excel":
            DataExporter.export_to_excel(records, filename)
        elif format == "json":
            DataExporter.export_to_json(records, filename)
        else:
            raise ValueError(f"Unsupported format: {format}")

        return filename


def export_history(history_manager,
                  format: str = "csv",
                  output_path: Optional[str] = None,
                  network: Optional[str] = None,
                  limit: Optional[int] = None) -> str:
    """
    Export history data with convenience function.

    Args:
        history_manager: GasHistory instance
        format: Export format ('csv', 'excel', or 'json')
        output_path: Path to save file (optional, auto-generated if None)
        network: Filter by network (optional)
        limit: Limit number of records (optional)

    Returns:
        Path to exported file
    """
    records = history_manager.get_records(network=network, limit=limit)

    if not records:
        raise ValueError("No records found to export")

    if output_path:
        if format == "csv":
            DataExporter.export_to_csv(records, output_path)
        elif format == "excel":
            DataExporter.export_to_excel(records, output_path)
        elif format == "json":
            DataExporter.export_to_json(records, output_path)
        else:
            raise ValueError(f"Unsupported format: {format}")
        return output_path
    else:
        return DataExporter.auto_export(records, format)
